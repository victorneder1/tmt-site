from __future__ import annotations

import os
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR / "data"))
BRUNO_COMPS_DIR = BASE_DIR / "tmt-site-bruno" / "projeto tabela comps"
DEFAULT_EXCEL_CANDIDATES = [
    DATA_DIR / "Big Telcos Data.xlsx",
    BASE_DIR / "Big Telcos Data.xlsx",
    BRUNO_COMPS_DIR / "Big Telcos Data.xlsx",
]
DEFAULT_VALUATION_CANDIDATES = [
    DATA_DIR / "telecom_comps_2026_2027.json",
    BASE_DIR / "telecom_comps_2026_2027.json",
    BRUNO_COMPS_DIR / "data.json",
]
DEFAULT_GLOBAL_COMPS_CANDIDATES = [
    DATA_DIR / "daily comps template 2025 - site telecom.xlsx",
    BASE_DIR / "daily comps template 2025 - site telecom.xlsx",
    BRUNO_COMPS_DIR / "daily comps template 2025 - site telecom.xlsx",
]

TIME_SERIES_SHEETS = [
    "Big telcos",
    "Service Revenue",
    "Mobile Service Revenue",
    "Fixed Revenue",
    "Net Revenue",
    "ARPU",
    "EBITDA",
    "Net Income",
    "Net Debt",
    "Capex",
    "OpFCF",
]

COMPANY_OPERATOR_MAP = {
    "Telefonica Brasil": "Vivo",
    "Vivo": "Vivo",
    "TIM Brasil": "TIM",
    "TIM": "TIM",
    "AMX": "AMX",
    "Claro": "Claro",
    "Oi S.A.": "Nio",
    "Oi": "Nio",
}

_CACHE: dict[str, Any] = {"path": None, "mtime": None, "payload": None}
CACHE_FILE = DATA_DIR / "telco_comps_cache.json"
CACHE_VERSION = 13


def invalidate_telco_comps_cache(remove_disk: bool = True) -> None:
    _CACHE.update({"path": None, "mtime": None, "payload": None})
    if remove_disk:
        try:
            CACHE_FILE.unlink(missing_ok=True)
        except Exception:
            pass


def _clean_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        text = value.strip()
        if text in {"", "#N/A", "#N/A N/A", "#VALUE!", "#NAME?"}:
            return None
        return text
    return value


def _is_period(value: Any) -> bool:
    if isinstance(value, (datetime, date, int)):
        if isinstance(value, int):
            return 1900 <= value <= 2100
        return True
    text = _clean_text(value)
    if not text:
        return False
    if text[:4].isdigit() and (len(text) == 4 or text[4:] in {"E", "Y"}):
        year = int(text[:4])
        return 1900 <= year <= 2100
    return bool(
        text[:2] in {"1Q", "2Q", "3Q", "4Q"}
        or text.endswith(("E", "Y"))
    )


def _resolve_excel_path() -> Path:
    explicit = os.environ.get("TELCOS_EXCEL_FILE")
    candidates = [Path(explicit)] if explicit else DEFAULT_EXCEL_CANDIDATES
    for path in candidates:
        if path.exists():
            return path
    checked = ", ".join(str(p) for p in candidates)
    raise FileNotFoundError(f"Big Telcos Data.xlsx not found. Checked: {checked}")


def _resolve_valuation_path() -> Path | None:
    explicit = os.environ.get("TELCOS_VALUATION_FILE")
    candidates = [Path(explicit)] if explicit else DEFAULT_VALUATION_CANDIDATES
    for path in candidates:
        if path.exists():
            return path
    return None


def _resolve_global_comps_path() -> Path | None:
    explicit = os.environ.get("TELCOS_GLOBAL_COMPS_FILE")
    candidates = [Path(explicit)] if explicit else DEFAULT_GLOBAL_COMPS_CANDIDATES
    for path in candidates:
        if path.exists():
            return path
    return None


def _to_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text in {"", "-", "n.a.", "N.A.", "n.m.", "N.M.", "#N/A", "#NUM!", "#VALUE!"}:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _load_valuation_table() -> dict[str, Any]:
    daily = _load_daily_local_valuation_table()
    if daily["companies"]:
        return daily

    path = _resolve_valuation_path()
    if not path:
        return {"source_file": None, "generated_at": None, "companies": []}

    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    companies = []
    for item in payload.get("companies", []):
        data = item.get("data", {})
        companies.append({
            "company": item.get("label"),
            "ticker": item.get("ticker"),
            "currency": item.get("currency"),
            "group": item.get("group"),
            "values": {
                year: {
                    "EV/EBITDA": year_data.get("ev_to_ebitda"),
                    "EV/Sales": year_data.get("ev_to_sales"),
                    "P/E": year_data.get("pe"),
                    "Dividend Yield": year_data.get("dividend_yield"),
                    "FCFE Yield": year_data.get("fcfe_yield"),
                    "EV/OpFCF": year_data.get("ev_to_opfcf"),
                }
                for year, year_data in data.items()
                if year in {"2026", "2027"}
            },
        })

    return {
        "source_file": path.name,
        "source_path": str(path),
        "last_modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
        "generated_at": payload.get("generated_at"),
        "companies": companies,
    }


def _load_daily_local_valuation_table() -> dict[str, Any]:
    path = _resolve_global_comps_path()
    if not path:
        return {"source_file": None, "generated_at": None, "companies": []}

    wb = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    try:
        if "Valuation" not in wb.sheetnames:
            return {"source_file": path.name, "source_path": str(path), "companies": []}
        ws = wb["Valuation"]
        company_map = {
            "Telefônica": "Telefonica Brasil",
            "Telefônica Brasil": "Telefonica Brasil",
            "TIM": "TIM Brasil",
            "Entel": "Entel",
            "AMX": "AMX",
            "Telecom Argentina": "TEO",
            "Unifique": "Unifique",
            "Brisanet": "Brisanet",
            "Desktop": "Desktop",
            "Megacable": "Megacable",
        }
        preferred_order = [
            "Telefonica Brasil",
            "TIM Brasil",
            "AMX",
            "Entel",
            "Megacable",
            "TEO",
            "Unifique",
            "Brisanet",
            "Desktop",
        ]
        currency_map = {
            "Telefonica Brasil": "R$",
            "TIM Brasil": "R$",
            "Unifique": "R$",
            "Brisanet": "R$",
            "Desktop": "R$",
            "AMX": "MXN",
            "Megacable": "MXN",
            "Entel": "CLP",
            "TEO": "US$",
        }
        by_company = {}
        for row_idx in range(5, ws.max_row + 1):
            raw_name = _clean_text(ws.cell(row_idx, 3).value).strip()
            raw_name = raw_name[:-1] if raw_name.endswith(" ") else raw_name
            company = company_map.get(raw_name)
            if not company:
                continue
            by_company[company] = {
                "company": company,
                "ticker": _clean_text(ws.cell(row_idx, 2).value),
                "currency": "",
                "market_cap": _to_number(ws.cell(row_idx, 16).value),
                "market_cap_currency": currency_map.get(company, ""),
                "group": "Daily Valuation",
                "values": {
                    "2026": {
                        "EV/Sales": _to_number(ws.cell(row_idx, 24).value),
                        "EV/EBITDA": _to_number(ws.cell(row_idx, 32).value),
                        "P/E": _to_number(ws.cell(row_idx, 48).value),
                        "Dividend Yield": _to_number(ws.cell(row_idx, 56).value),
                        "FCFE Yield": _to_number(ws.cell(row_idx, 64).value),
                        "EV/OpFCF": _to_number(ws.cell(row_idx, 40).value),
                    },
                    "2027": {
                        "EV/Sales": _to_number(ws.cell(row_idx, 25).value),
                        "EV/EBITDA": _to_number(ws.cell(row_idx, 33).value),
                        "P/E": _to_number(ws.cell(row_idx, 49).value),
                        "Dividend Yield": _to_number(ws.cell(row_idx, 57).value),
                        "FCFE Yield": _to_number(ws.cell(row_idx, 65).value),
                        "EV/OpFCF": _to_number(ws.cell(row_idx, 41).value),
                    },
                },
            }

        return {
            "source_file": path.name,
            "source_path": str(path),
            "last_modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            "generated_at": None,
            "companies": [by_company[name] for name in preferred_order if name in by_company],
        }
    finally:
        wb.close()


def _load_global_telecom_comps() -> dict[str, Any]:
    path = _resolve_global_comps_path()
    if not path:
        return {"source_file": None, "generated_at": None, "companies": []}

    wb = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    try:
        if "Comps" not in wb.sheetnames:
            return {
                "source_file": path.name,
                "source_path": str(path),
                "last_modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                "companies": [],
            }
        ws = wb["Comps"]
        companies = []
        by_company: dict[str, dict[str, Any]] = {}
        for row_idx in range(7, ws.max_row + 1):
            company = _clean_text(ws.cell(row_idx, 2).value)
            if not company:
                continue
            if company.lower() == "median":
                break
            item = {
                "company": company,
                "ticker": _clean_text(ws.cell(row_idx, 3).value),
                "currency": "",
                "market_cap": _to_number(ws.cell(row_idx, 6).value),
                "market_cap_currency": "US$",
                "group": "Global Telcos",
                "values": {
                    "2026": {
                        "EV/Sales": _to_number(ws.cell(row_idx, 13).value),
                        "EV/EBITDA": _to_number(ws.cell(row_idx, 16).value),
                        "P/E": _to_number(ws.cell(row_idx, 9).value),
                        "Dividend Yield": _to_number(ws.cell(row_idx, 28).value),
                        "FCFE Yield": _to_number(ws.cell(row_idx, 31).value),
                        "EV/OpFCF": _to_number(ws.cell(row_idx, 43).value),
                    },
                    "2027": {
                        "EV/Sales": _to_number(ws.cell(row_idx, 14).value),
                        "EV/EBITDA": _to_number(ws.cell(row_idx, 17).value),
                        "P/E": _to_number(ws.cell(row_idx, 10).value),
                        "Dividend Yield": _to_number(ws.cell(row_idx, 29).value),
                        "FCFE Yield": _to_number(ws.cell(row_idx, 32).value),
                        "EV/OpFCF": _to_number(ws.cell(row_idx, 44).value),
                    },
                },
            }
            key = company.strip().lower()
            current = by_company.get(key)
            if (
                current is None
                or (_to_number(item.get("market_cap")) or 0) > (_to_number(current.get("market_cap")) or 0)
            ):
                by_company[key] = item
        companies = list(by_company.values())

        return {
            "source_file": path.name,
            "source_path": str(path),
            "last_modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            "companies": companies,
        }
    finally:
        wb.close()


def _load_disk_cache(excel_path: Path, excel_mtime: float, valuation_path: Path | None, global_comps_path: Path | None) -> dict[str, Any] | None:
    if not CACHE_FILE.exists():
        return None
    try:
        payload = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None

    meta = payload.get("_cache_meta", {})
    if meta.get("version") != CACHE_VERSION:
        return None
    if meta.get("excel_path") != str(excel_path) or meta.get("excel_mtime") != excel_mtime:
        return None

    valuation_mtime = valuation_path.stat().st_mtime if valuation_path and valuation_path.exists() else None
    if meta.get("valuation_path") != (str(valuation_path) if valuation_path else None):
        return None
    if meta.get("valuation_mtime") != valuation_mtime:
        return None

    global_comps_mtime = global_comps_path.stat().st_mtime if global_comps_path and global_comps_path.exists() else None
    if meta.get("global_comps_path") != (str(global_comps_path) if global_comps_path else None):
        return None
    if meta.get("global_comps_mtime") != global_comps_mtime:
        return None

    payload.pop("_cache_meta", None)
    return payload


def _load_stale_disk_cache(error: Exception) -> dict[str, Any] | None:
    if not CACHE_FILE.exists():
        return None
    try:
        payload = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None
    meta = payload.pop("_cache_meta", {})
    payload["cache_warning"] = (
        "Using the last cached telecom comps payload because the current Big Telcos workbook "
        f"could not be read: {type(error).__name__}: {error}"
    )
    payload["cache_meta"] = {
        "stale": True,
        "cached_excel_path": meta.get("excel_path"),
        "cached_excel_mtime": meta.get("excel_mtime"),
    }
    return _normalize_payload(payload)


def _write_disk_cache(payload: dict[str, Any], excel_path: Path, excel_mtime: float, valuation_path: Path | None, global_comps_path: Path | None) -> None:
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        cache_payload = dict(payload)
        cache_payload["_cache_meta"] = {
            "version": CACHE_VERSION,
            "excel_path": str(excel_path),
            "excel_mtime": excel_mtime,
            "valuation_path": str(valuation_path) if valuation_path else None,
            "valuation_mtime": valuation_path.stat().st_mtime if valuation_path and valuation_path.exists() else None,
            "global_comps_path": str(global_comps_path) if global_comps_path else None,
            "global_comps_mtime": global_comps_path.stat().st_mtime if global_comps_path and global_comps_path.exists() else None,
        }
        temp_path = CACHE_FILE.with_suffix(".tmp")
        temp_path.write_text(json.dumps(cache_payload, ensure_ascii=False), encoding="utf-8")
        temp_path.replace(CACHE_FILE)
    except Exception:
        pass


def _parse_summary(ws) -> dict[str, Any]:
    header_row = None
    company_col = None
    for row in ws.iter_rows():
        for cell in row:
            if _clean_text(cell.value).lower() == "company":
                header_row = cell.row
                company_col = cell.column
                break
        if header_row:
            break

    if not header_row or not company_col:
        return {"columns": [], "rows": []}

    year_row = header_row + 1
    columns = []
    current_metric = ""
    for col in range(company_col + 1, ws.max_column + 1):
        metric = _clean_text(ws.cell(header_row, col).value)
        if metric:
            current_metric = metric.replace(" / ", "/").replace("EV/ OpFCF", "EV/OpFCF")
        year = _clean_text(ws.cell(year_row, col).value)
        if current_metric and year:
            columns.append({"col": col, "metric": current_metric, "year": year})

    rows = []
    for row_idx in range(year_row + 1, ws.max_row + 1):
        company = _clean_text(ws.cell(row_idx, company_col).value)
        if not company:
            continue
        values = {}
        has_value = False
        for col_meta in columns:
            value = _clean_value(ws.cell(row_idx, col_meta["col"]).value)
            key = f'{col_meta["metric"]}|{col_meta["year"]}'
            values[key] = value
            has_value = has_value or value is not None
        if has_value:
            rows.append({
                "company": company,
                "operator": COMPANY_OPERATOR_MAP.get(company),
                "is_benchmark": company.lower() in {
                    "median", "telecom integrated", "us median", "european median", "asian median"
                },
                "values": values,
            })

    return {
        "columns": [{"metric": c["metric"], "year": c["year"], "key": f'{c["metric"]}|{c["year"]}'} for c in columns],
        "rows": rows,
    }


def _normalize_section_company(sheet_name: str, title: str, company: str, label: str) -> str:
    # The workbook currently labels Claro Brazil's nominal mobile service revenue row as AMX,
    # while the growth row is correctly named Claro. Keep America Movil separate elsewhere.
    if sheet_name == "Mobile Service Revenue" and title == "Mobile Service Revenue" and company == "AMX":
        return "Claro"
    return company


def _normalize_series_values(sheet_name: str, title: str, company: str, values: list[Any]) -> list[Any]:
    # Unifique's net income row is supplied in thousands while the other ISP
    # financial rows are in millions. Normalize it to match the chart scale.
    if sheet_name == "Net Income" and title == "Net Income" and company == "Unifique":
        return [
            value / 1000 if isinstance(value, (int, float)) and abs(value) > 1000 else value
            for value in values
        ]
    return values


def _normalize_payload(payload: dict[str, Any]) -> dict[str, Any]:
    for sheet_name, sections in payload.get("sheets", {}).items():
        for section in sections:
            title = section.get("title", "")
            for serie in section.get("series", []):
                serie["values"] = _normalize_series_values(
                    sheet_name,
                    title,
                    serie.get("company", ""),
                    serie.get("values", []),
                )
    return payload


def _parse_sheet_sections(ws, sheet_name: str) -> list[dict[str, Any]]:
    sections = []
    row_idx = 1
    while row_idx <= ws.max_row:
        title = _clean_text(ws.cell(row_idx, 2).value)
        periods = []
        seen_period_labels = set()
        for col in range(3, ws.max_column + 1):
            raw = ws.cell(row_idx, col).value
            if _is_period(raw):
                label = _clean_value(raw)
                key = str(label)
                if key in seen_period_labels:
                    continue
                seen_period_labels.add(key)
                periods.append({"col": col, "label": label})
        if title and len(periods) >= 2:
            series = []
            scan = row_idx + 1
            while scan <= ws.max_row:
                company = _clean_text(ws.cell(scan, 2).value)
                next_title_periods = [
                    ws.cell(scan, col).value for col in range(3, min(ws.max_column, 8) + 1)
                ]
                if company and any(_is_period(v) for v in next_title_periods):
                    break
                if not company:
                    if series:
                        break
                    scan += 1
                    continue
                values = [_clean_value(ws.cell(scan, p["col"]).value) for p in periods]
                if any(v is not None for v in values):
                    label = _clean_text(ws.cell(scan, 1).value)
                    company_name = company.replace(" y/y growth", "")
                    normalized_company = _normalize_section_company(sheet_name, title, company_name, label)
                    series.append({
                        "company": normalized_company,
                        "label": label,
                        "values": _normalize_series_values(sheet_name, title, normalized_company, values),
                    })
                scan += 1
            if series:
                sections.append({
                    "title": title,
                    "periods": [p["label"] for p in periods],
                    "series": series,
                })
            row_idx = max(scan, row_idx + 1)
        else:
            row_idx += 1
    return sections


def _parse_company_cash_flow(ws) -> dict[str, Any] | None:
    for row_idx in range(1, ws.max_row + 1):
        if _clean_text(ws.cell(row_idx, 1).value).lower() != "cash flow":
            continue
        periods = []
        for col in range(2, ws.max_column + 1):
            raw = ws.cell(row_idx, col).value
            if _is_period(raw):
                periods.append({"col": col, "label": _clean_value(raw)})
        if not periods:
            return None
        series = []
        scan = row_idx + 1
        while scan <= ws.max_row:
            label = _clean_text(ws.cell(scan, 1).value)
            if not label:
                break
            values = [_clean_value(ws.cell(scan, p["col"]).value) for p in periods]
            if any(v is not None for v in values):
                series.append({"label": label, "values": values})
            scan += 1
        return {"periods": [p["label"] for p in periods], "series": series}
    return None


def load_telco_comps() -> dict[str, Any]:
    path = _resolve_excel_path()
    mtime = path.stat().st_mtime
    valuation_path = _resolve_valuation_path()
    global_comps_path = _resolve_global_comps_path()
    global_mtime = global_comps_path.stat().st_mtime if global_comps_path and global_comps_path.exists() else None
    cache_key = f"{path}|{valuation_path}|{global_comps_path}"
    if _CACHE["path"] == cache_key and _CACHE["mtime"] == (mtime, global_mtime) and _CACHE["payload"] is not None:
        return _CACHE["payload"]

    cached = _load_disk_cache(path, mtime, valuation_path, global_comps_path)
    if cached is not None:
        cached = _normalize_payload(cached)
        _CACHE.update({"path": cache_key, "mtime": (mtime, global_mtime), "payload": cached})
        return cached

    try:
        wb = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    except Exception as exc:
        stale = _load_stale_disk_cache(exc)
        if stale is not None:
            _CACHE.update({"path": cache_key, "mtime": (mtime, global_mtime), "payload": stale})
            return stale
        raise
    sheets = {}
    for sheet_name in TIME_SERIES_SHEETS:
        if sheet_name in wb.sheetnames:
            sheets[sheet_name] = _parse_sheet_sections(wb[sheet_name], sheet_name)

    payload = {
        "source_file": path.name,
        "source_path": str(path),
        "last_modified": datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M"),
        "company_operator_map": COMPANY_OPERATOR_MAP,
        "summary": _parse_summary(wb["Summary Tele.Comps"]) if "Summary Tele.Comps" in wb.sheetnames else {"columns": [], "rows": []},
        "valuation_2026_2027": _load_valuation_table(),
        "global_telecom_comps": _load_global_telecom_comps(),
        "company_financials": {
            "Vivo": {
                "cash_flow": _parse_company_cash_flow(wb["Vivo"]) if "Vivo" in wb.sheetnames else None,
            },
        },
        "sheets": sheets,
    }
    payload = _normalize_payload(payload)
    wb.close()
    _write_disk_cache(payload, path, mtime, valuation_path, global_comps_path)
    _CACHE.update({"path": cache_key, "mtime": (mtime, global_mtime), "payload": payload})
    return payload
