from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook


DOCUMENT_HEADERS = [
    "empresa",
    "tipo_documento",
    "nome_companhia",
    "categoria",
    "tipo",
    "data_referencia",
    "data_entrega",
    "protocolo",
    "versao",
    "parse_status",
    "parse_error",
    "buyback_count",
    "movement_count",
    "link_download",
]

MOVEMENT_HEADERS = [
    "empresa",
    "tipo_documento",
    "protocolo",
    "holder_name",
    "holder_role",
    "holder_group",
    "asset",
    "title_characteristics",
    "intermediary",
    "operation_type",
    "operation_day",
    "quantity",
    "price_avg",
    "financial_volume",
    "initial_quantity",
    "final_quantity",
    "no_operations",
    "is_buyback",
    "details",
    "reference_date",
    "delivery_date",
]


def export_workbook(
    documents: list[dict[str, Any]],
    movements: list[dict[str, Any]],
    workbook_path: Path,
    parsed_data_path: Path,
) -> Path:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    parsed_data_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    documents_sheet = workbook.active
    documents_sheet.title = "documentos"
    documents_sheet.append(DOCUMENT_HEADERS)
    for document in documents:
        summary = document.get("summary", {})
        documents_sheet.append(
            [
                document.get("company_alias", ""),
                document.get("document_kind", ""),
                document.get("Nome_Companhia", ""),
                document.get("Categoria", ""),
                document.get("Tipo", ""),
                document.get("Data_Referencia", ""),
                document.get("Data_Entrega", ""),
                document.get("Protocolo_Entrega", ""),
                document.get("Versao", ""),
                document.get("parse_status", ""),
                document.get("parse_error", ""),
                int(summary.get("buyback_count", 0) or 0),
                int(summary.get("movement_count", 0) or 0),
                document.get("Link_Download", ""),
            ]
        )

    movement_sheet = workbook.create_sheet("movimentacoes")
    movement_sheet.append(MOVEMENT_HEADERS)
    for movement in movements:
        movement_sheet.append([movement.get(header, "") for header in MOVEMENT_HEADERS])

    summary_sheet = workbook.create_sheet("resumo_mensal")
    summary_sheet.append(
        [
            "empresa",
            "ano_entrega",
            "mes_entrega",
            "ano_referencia",
            "mes_referencia",
            "tipo_documento",
            "documentos",
            "movimentacoes",
            "movimentacoes_com_quantidade",
            "quantidade_total",
            "volume_financeiro_total",
            "eventos_recompra",
        ]
    )
    for row in build_monthly_summary(documents, movements):
        summary_sheet.append(row)

    try:
        workbook.save(workbook_path)
        final_workbook_path = workbook_path
    except PermissionError:
        final_workbook_path = workbook_path.with_name(f"{workbook_path.stem}.updated{workbook_path.suffix}")
        workbook.save(final_workbook_path)
    parsed_data_path.write_text(
        json.dumps({"documents": documents, "movements": movements}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return final_workbook_path


def build_monthly_summary(
    documents: list[dict[str, Any]],
    movements: list[dict[str, Any]],
) -> list[list[Any]]:
    document_index: dict[tuple[str, str, str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "documents": set(),
            "movements": 0,
            "movements_with_quantity": 0,
            "quantity_total": 0.0,
            "financial_volume_total": 0.0,
            "buyback_events": 0,
            "reference_year": "",
            "reference_month": "",
        }
    )
    for document in documents:
        delivery_date = document.get("Data_Entrega", "")
        reference_date = document.get("Data_Referencia", "")
        if len(delivery_date) < 7:
            continue
        year, month = delivery_date[:4], delivery_date[5:7]
        key = (document.get("company_alias", ""), year, month, document.get("document_kind", ""))
        document_index[key]["documents"].add(document.get("Protocolo_Entrega", ""))
        if len(reference_date) >= 7:
            document_index[key]["reference_year"] = reference_date[:4]
            document_index[key]["reference_month"] = reference_date[5:7]

    for movement in movements:
        delivery_date = str(movement.get("delivery_date", ""))
        reference_date = str(movement.get("reference_date", ""))
        if len(delivery_date) < 7:
            continue
        year, month = delivery_date[:4], delivery_date[5:7]
        key = (movement.get("company_alias", ""), year, month, movement.get("document_kind", ""))
        item = document_index[key]
        item["movements"] += 1
        if movement.get("quantity") is not None:
            item["movements_with_quantity"] += 1
            item["quantity_total"] += float(movement.get("quantity") or 0)
        item["financial_volume_total"] += float(movement.get("financial_volume") or 0)
        item["buyback_events"] += int(movement.get("is_buyback") or 0)
        if len(reference_date) >= 7 and not item["reference_year"]:
            item["reference_year"] = reference_date[:4]
            item["reference_month"] = reference_date[5:7]

    rows: list[list[Any]] = []
    for key in sorted(document_index.keys(), reverse=True):
        company_alias, year, month, document_kind = key
        item = document_index[key]
        rows.append(
            [
                company_alias,
                int(year),
                int(month),
                int(item["reference_year"]) if item["reference_year"] else "",
                int(item["reference_month"]) if item["reference_month"] else "",
                document_kind,
                len(item["documents"]),
                item["movements"],
                item["movements_with_quantity"],
                round(item["quantity_total"], 5),
                round(item["financial_volume_total"], 5),
                item["buyback_events"],
            ]
        )
    return rows
